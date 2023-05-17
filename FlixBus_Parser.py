import requests
import openpyxl
import pprint
import subprocess

    
list_of_cities = [ "Düsseldorf", "Berlin", "Dortmund", "Cologne", "Köln", "Essen", "Stuttgart", "München"]
city_name_Arv = "Vaduz"
data_of_travel = "19.05.2023"
 
######
## Preparing Excel
######
# Create a new workbook
workbook = openpyxl.Workbook()

# Select the active sheet
sheet = workbook.active
row_id = 1
## Creating the headers in excel
Headers_excel = ["City", "Departure Stop", "Departure Date", "Arrival Stop", "Arrival Date", "Total Time", "Price", "Seats Left"]
# Loop through the list and store information in Excel
for i, data in enumerate(Headers_excel, start=1):
    sheet.cell(row=row_id, column=i, value=data)


## Parsing through each city
row_id = 2
for city_name_Dep in list_of_cities:
    error = 0
    city_id_Dep = "";
    city_id_Arv = "";
    parsed_Data = []
    
    url = 'https://global.api.flixbus.com/search/autocomplete/cities?q='+ city_name_Dep +'&lang=de&country=de&flixbus_cities_only=false'

    # Make a GET request to the API endpoint to search for cities
    response = requests.get(f"{url}")
    # Extract the JSON data from the response
    data = response.json()

    # Loop through the data to find the city with the matching name
    for city in data:
        if city["name"].lower() == city_name_Dep.lower():
            city_id_Dep = city["id"]
            #print(f"The unique ID for {city_name_Dep} is {city_id_Dep}")
            break
    else:
        print(f"No city found with the name {city_name_Dep}")
        parsed_Data.append(city_name_Dep + " Not Found")
        error = 1




    url = 'https://global.api.flixbus.com/search/autocomplete/cities?q='+ city_name_Arv +'&lang=de&country=de&flixbus_cities_only=false'

    # Make a GET request to the API endpoint to search for cities
    response = requests.get(f"{url}")
    # Extract the JSON data from the response
    data = response.json()

    # Loop through the data to find the city with the matching name
    for city in data:
        if city["name"].lower() == city_name_Arv.lower():
            city_id_Arv = city["id"]
            #print(f"The unique ID for {city_name_Arv} is {city_id_Arv}")
            break
    else:
        print(f"No city found with the name {city_name_Arv}")
        parsed_Data.append(city_name_Arv + " Not Found")
        error = 1
        
        
                

    url = 'https://global.api.flixbus.com/search/service/v4/search?from_city_id='+city_id_Dep+'&to_city_id='+city_id_Arv+'&departure_date='+data_of_travel+'&products=%7B%22adult%22%3A1%7D&currency=EUR&locale=de&search_by=cities&include_after_midnight_rides=1'

    # Make a GET request to the API endpoint to search for cities
    response = requests.get(f"{url}")
    # Extract the JSON data from the response
    data = response.json()
    #pprint.pprint(data)


    if(error == 0):
        Station_names = [city['name'] for city in data['stations'].values()]
        Station_ids = [city['id'] for city in data['stations'].values()]

        Station_id_to_name = {id: name for id, name in zip(Station_ids, Station_names)}

        for trip in data['trips']:
           #duration_id = trip['departure']['duration']['id']
           for res in trip['results']:
            print("\nTrip", row_id -1)     
            print("Departure Station",Station_id_to_name.get(trip['results'][res]['arrival']['station_id']))  
            print("Departure Date",trip['results'][res]['departure']['date'])
            print("Arrival Date",trip['results'][res]['arrival']['date'])
            print("Arrival Station",Station_id_to_name.get(trip['results'][res]['arrival']['station_id']))                 
            print("Price",trip['results'][res]['price']['total'])
            
            
            parsed_Data.append(city_name_Dep)
            parsed_Data.append(Station_id_to_name.get(trip['results'][res]['departure']['station_id']))
            parsed_Data.append(trip['results'][res]['departure']['date'])
            parsed_Data.append(Station_id_to_name.get(trip['results'][res]['arrival']['station_id']))
            parsed_Data.append(trip['results'][res]['arrival']['date'])
            parsed_Data.append(str(trip['results'][res]['duration']['hours']) + ":" + str(trip['results'][res]['duration']['minutes']))
            parsed_Data.append(trip['results'][res]['price']['total'])
            parsed_Data.append(trip['results'][res]['available']['seats'])
            
            # Store information in Excel
            for i, data in enumerate(parsed_Data, start=1):
                sheet.cell(row=row_id, column=i, value=data)
            row_id = row_id + 1
            parsed_Data = []  
    else:
        parsed_Data.append(city_name_Dep)
        parsed_Data.append("No trips found")
        
        
        # Store information in Excel
        for i, data in enumerate(parsed_Data, start=1):
            sheet.cell(row=row_id, column=i, value=data)
        row_id = row_id + 1
        parsed_Data = []    
# Save the workbook
file_path = 'FlexBus_Planner.xlsx'
workbook.save(file_path)
# subprocess.call(['open', file_path])